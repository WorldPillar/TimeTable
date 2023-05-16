import openpyxl
from openpyxl.styles import Alignment, Font

import utils
from schooldata.data import SchoolData
from schooldata.school import School, Lesson


def _get_len(school: School, param: str) -> int:
    """
    Возвращает длину списка student_classes при param 'class'
     или длину списка teachers при param 'teacher' у объекта school
    :param school: Объект класса School
    :param param: 'class' или 'teacher'
    """
    if param == 'class':
        return len(school.student_classes)
    else:
        return len(school.teachers)


def _get_name(school: School, pos: int, param: str) -> str:
    """
    Возвращает имя объекта списка student_classes при param 'class'
     или имя объекта списка teachers при param 'teacher' у объекта school
    :param school: Объект класса School
    :param pos: Позиция объекта в списке
    :param param: 'class' или 'teacher'
    """
    if param == 'class':
        return school.student_classes[pos].get_name()
    else:
        return school.teachers[pos].get_name()


def _get_id(lesson: Lesson, param: str) -> [int]:
    """
    Возвращает id объекта student_class при param 'class'
     или id объекта teacher при param 'teacher' у объекта lesson
    :param lesson: Объект класса Lesson
    :param param: 'class' или 'teacher'
    """
    if param == 'class':
        return [lesson.student_class.id]
    else:
        return [teacher.id for teacher in lesson.teachers]


class ExcelProcessor:
    @staticmethod
    def export_table(school: School, param: str, rout: str) -> bool:
        """
        Возвращает результат экспорта в виде bool.
        :param school: Объект класса School, в котором хранятся все данные для экспорта.
        :param param: Параметр 'class' для экспорта расписания классов; Параметр 'teacher' для экспорта расписания учителей.
        :param rout: Путь к файлу записи.
        """

        wb = openpyxl.Workbook()
        ws = wb.active
        # Заполняем заголовки
        for i in range(school.amount_lessons * school.amount_days):
            day = i // school.amount_lessons
            if i % school.amount_lessons == 0:
                day_cell = ws.cell(i + 2, 1)
                day_cell.value = SchoolData.get_day_name(day)
                day_cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=90, )
                day_cell.font = Font(bold=True)

                ws.merge_cells(start_row=i + 2, start_column=1, end_row=i + school.amount_lessons + 1, end_column=1)
            position_cell = ws.cell(i + 2, 2)
            position_cell.alignment = Alignment(horizontal='center', vertical='center')
            position_cell.value = i % school.amount_lessons + 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        for i in range(_get_len(school, param)):
            cell = ws.cell(1, i + 3)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.value = _get_name(school, i, param)
            for j in range(school.amount_lessons * school.amount_days):
                day, les_pos = utils.column_to_days_lessons(j, school.amount_lessons)
                for les in school.timetable[day][les_pos]:
                    ids = _get_id(les, param)
                    for one_id in ids:
                        lesson_cell = ws.cell(j + 2, one_id + 3)
                        lesson_cell.alignment = Alignment(horizontal='center', vertical='center')

                        output = les.student_class.name if param == 'teacher' else les.subject.abbreviation
                        lesson_cell.value = output
        try:
            wb.save(rout)
            wb.close()
        except IOError:
            return False
        return True
